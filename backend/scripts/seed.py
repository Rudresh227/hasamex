import openpyxl
import asyncio
import sys
import os
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

# Add parent directory to path so we can import 'app'
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import async_session
from app.models.expert import Expert, ExpertRate, ExpertFile, LookupValue
import uuid

# Map Excel headers to our model fields
COLUMN_MAP = {
    'Expert ID': 'expert_id',
    'First Name': 'first_name',
    'Last Name': 'last_name',
    'Primary Email 1': 'primary_email',
    'Secondary Email': 'secondary_email',
    'LinkedIn URL': 'linkedin_url',
    'Location': 'location',
    'Timezone': 'timezone',
    'Region': 'region',
    'Current Employment Status': 'employment_status',
    'Seniority': 'seniority',
    'Years of Experience': 'years_experience',
    'Title/Headline': 'headline',
    'Bio': 'bio',
    'Strength Topics': 'strength_topics',
    'Primary Sector': 'sector',
    'Company Role': 'company_role',
    'Expert Function': 'function',
    'HCMS Classification': 'hcms_class',
    'Expert Status': 'expert_status',
    'Payment Details': 'payment_details',
    'Notes': 'notes',
    'Events Invited To': 'events_invited',
    'Total Calls Completed': 'total_calls',
}

async def get_lookup_id(db, category, value):
    if not value or value == "-":
        return None
    stmt = select(LookupValue.id).where(LookupValue.category == category, LookupValue.value == value)
    result = await db.execute(stmt)
    return result.scalar_one_or_none()

async def seed_from_excel(file_path: str):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    # The plan says 'Current DB' sheet, but let's check what's actually there if needed
    ws = wb.active # Usually the first sheet
    
    headers = [c.value for c in ws[1]]
    
    async with async_session() as db:
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {COLUMN_MAP[h]: v for h, v in zip(headers, row) if h in COLUMN_MAP}
            
            if not data.get('expert_id'):
                continue

            # Check if exists
            stmt = select(Expert).where(Expert.expert_id == data['expert_id'])
            result = await db.execute(stmt)
            if result.scalar_one_or_none():
                print(f"Skipping {data['expert_id']}, already exists.")
                continue

            # Map lookups
            region_id = await get_lookup_id(db, 'region', data.get('region'))
            emp_status_id = await get_lookup_id(db, 'employment_status', data.get('employment_status'))
            sector_id = await get_lookup_id(db, 'sector', data.get('sector'))
            func_id = await get_lookup_id(db, 'function', data.get('function'))
            exp_status_id = await get_lookup_id(db, 'expert_status', data.get('expert_status'))

            expert_data = {
                k: v for k, v in data.items() 
                if k not in ['region', 'employment_status', 'sector', 'function', 'expert_status']
            }
            
            db_expert = Expert(
                **expert_data,
                region_id=region_id,
                employment_status_id=emp_status_id,
                sector_id=sector_id,
                function_id=func_id,
                expert_status_id=exp_status_id
            )
            
            # Handle rate if present in a messy way (simplified for now)
            # Find the Rate column if it exists as 'Hourly Rate' and 'Currency'
            # Assuming headers contain them
            currency = None
            rate = None
            for h, v in zip(headers, row):
                if h == 'Currency': currency = v
                if h == 'Hourly Rate': rate = v
            
            if currency and rate and rate != "-":
                try:
                    db_expert.rates.append(ExpertRate(currency=currency, hourly_rate=float(rate), is_primary=True))
                except (ValueError, TypeError):
                    pass

            db.add(db_expert)
            print(f"Added {data['expert_id']}: {data['first_name']} {data['last_name']}")
        
        await db.commit()
        print("Seeding complete.")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "..", "..", "Documents", "Test Hasamex Expert Database (HEB).xlsx")
    asyncio.run(seed_from_excel(excel_path))
